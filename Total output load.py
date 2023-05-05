import json
import openpyxl
import sys
from openpyxl import Workbook
def build_file(address, aim_file):
    file_address = address
    trans_file = json.dumps(aim_file)
    with open(file_address, 'w') as json_build_object:
        json_build_object.write(trans_file)

def load_file(address):
    file_address = address
    with open(file_address) as add_json:
        new_dict = json.load(add_json)
    return new_dict

port_dict = load_file("port max load.json")

port_total_load = {}

for port in port_dict.keys():
    port_total_load[port] = [0,0]#离开（装货），到达（卸货）

whole_od_path = load_file("whole od path (key as number).json")
for key in whole_od_path.keys():
    for path in whole_od_path[key]:
        for part in path:
            d_port = part[0]
            a_port = part[1]
            port_total_load[d_port][0] += 1
            port_total_load[a_port][1] -= 1

aim_file_address = "port total load.xlsx"
wb = Workbook()
ws1 = wb.active
ws1.cell(row=1,column=1,value="port")
ws1.cell(row=1,column=2,value="depart(load load)")
ws1.cell(row=1,column=3,value="arrive(unload load)")
r = 2
for p in port_total_load:
    ws1.cell(row=r,column=1,value=p)
    ws1.cell(row=r,column=2,value=port_total_load[p][0])
    ws1.cell(row=r,column=3,value=port_total_load[p][1])
    r += 1

wb.save(aim_file_address)