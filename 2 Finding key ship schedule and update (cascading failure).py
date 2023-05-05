# 关键的船期属于特定的航线
# 抽离出特定的航线
# 匹配对应的时间（中断是如何影响船期的，应该如何变动？）
# 在计算最短路的基础上，统计各个节点区间的访问次数
# coding:utf-8
# 找到路线中中断的两个航程
# 记录航线中出现suez_canal的位置
from typing import List, Dict, Any


def key_port_address(line_list, key_port):  # 找到关键港口所在的列表的位置
    port_address = []
    p_1 = line_list.index(key_port)
    port_address.append(p_1)
    p_2 = line_list.index(key_port, p_1 + 1, len(line_list) + 1)
    port_address.append(p_2)
    return port_address


# 在计算最短路的基础上，统计各个节点区间的访问次数
# coding:utf-8
import openpyxl  # 导入工作蒲模块
from openpyxl.utils import get_column_letter  # 导入模块中读取最大列的函数
from os import listdir
import os
import datetime
# 设置的变量务必不可以跟模块、函数等重名！！！
import time
import copy
import calendar
import json
import math
import sys
import gc

start_time = time.time()
handing_time = datetime.timedelta(days=1)
half_hour = datetime.timedelta(seconds=1800)
# ===================================================================================================================
# 为算法加载初步的数据
# 将文件夹中的所有工作簿的工作表以邻接、直达的路径时间导入短期边中，边包含所属航线信息
open_file_path = r'E:\大连海事大学\博士\数据\集装箱船公司船期数据\中远海集运（按照区域航线划分）\标准化数据'

file_name = listdir(open_file_path)
c_block = ['AET Asia Europe.xlsx', 'TPT.xlsx']

true_port_list = []  # type: List[str] # 永久列表
new_od_dict = {} # 记录节点对路径的字典，全局变量
load_dict = {}  # 储存时刻使用次数的统计容器，全局变量

p_list = []  # 临时表
time_list = []  # 临时表
date_list = {}  # 临时表日期表换成字典储存
datetime_list = {}  # 临时表日期时间变更为字典储存

def load_file(address):
    file_address = address
    with open(file_address) as add_json:
        new_dict = json.load(add_json)
    return new_dict

def build_file(address, aim_file):
    file_address = address
    trans_file = json.dumps(aim_file)
    with open(file_address, 'w') as json_build_object:
        json_build_object.write(trans_file)

def recover_whole_od(aim_dict):  # 复原初始的od路径json文件
    new_dict = {}
    for key in aim_dict.keys():
        try:
            new_key = (aim_dict[key][0][0][0], aim_dict[key][0][-1][1])
        except IndexError:
            print("key" + str(key))
            print(str(aim_dict[key]))
            sys.exit()
        for key_path in aim_dict[key]:
            for path_part in key_path:
                l_time = datetime.datetime.strptime(path_part[2], "%Y-%m-%d %H:%M")
                a_time = datetime.datetime.strptime(path_part[3], "%Y-%m-%d %H:%M")
                path_part[2] = l_time
                path_part[3] = a_time
        new_dict[new_key] = copy.deepcopy(aim_dict[key])
    return new_dict

def disturb_time(s_time, e_time):
    disturb_list = []
    stamp_time1 = datetime.datetime.strptime(s_time, "%Y-%m-%d %H:%M")
    stamp_time2 = datetime.datetime.strptime(e_time, "%Y-%m-%d %H:%M")
    disturb_list.append(stamp_time1)
    disturb_list.append(stamp_time2)
    return disturb_list


# 加入中断时间
break_name = "中断1天"
disturb_interval = disturb_time("2020-3-23 00:00", "2020-3-24 00:00")
k_port = "SUEZ CANAL"
trans_canal = datetime.timedelta(hours=12)
delay_voyage = []  # 用于记录延长的船期
schedule_dict = {}
break_impact = []
overload_way = {}  # type: Dict[Any, Any]
# last_overload = []
for i in file_name:
    com_path = os.path.join(open_file_path, i)
    wb = openpyxl.load_workbook(com_path)
    sheet_names = wb.sheetnames  # 获取工作蒲中所有表格
    for j in sheet_names:  # 从表格中提取港口名、时刻、日期存入三个临时列表（重置）
        strs = str(j)  # 按照航线名称创建字典键
        sheet = wb[j]
        max_column = sheet.max_column
        max_row = sheet.max_row  # 获取最大行数
        column = get_column_letter(max_column)
        row_1 = sheet["I3":"%s3" % column]
        for row1_cells in row_1:
            for cell in row1_cells:
                p_list.append(str(cell.value))  # 读取出每个工作表的港口信息
        schedule_dict[strs] = copy.deepcopy(p_list)
        row_2 = sheet["I5":"%s5" % column]
        for row2_cells in row_2:
            for cell in row2_cells:
                time_list.append(str(cell.value))
        for row_number in range(6, max_row + 1):  # 遍历行，对row3做修正==修正！！！！成立一个关日期行的字典？后期添加航线字典也使用字典值的遍历！！！！！
            first_column = str("I" + str(row_number))
            last_column = str("%s" + str(row_number))
            row_3 = sheet[first_column:last_column % column]  # row_3 = sheet["I6":"%s6" % column]
            voyage_number = len(date_list)
            date_list[voyage_number] = []
            for row3_cells in row_3:
                for cell in row3_cells:
                    date_list[voyage_number].append(str(cell.value))
        for port in p_list:  # 生成真实的港口列表
            if port not in true_port_list:
                if port != 'None':
                    true_port_list.append(port)
                    load_dict[port] = {}
                    # 计划设计{}中的数据为time：[次数，装卸货时间]当前是到达时间，则是卸货，前时间-后时间符号为负的；当前是离开时间，则是装货，后时间-前时间，符号为正的
                    # 不确定这样的赋值方式是否适用！！！！！！！！！！
        n = len(p_list)
        # print(str(n))
        # 在检查结果的时候，出现"None 13:00"的字样，原因是有航次在这个日期没有船舶靠泊的安排，但是不能将其从列表中删除，否则出现列表串位的情况
        # 因此在发现有None的时候，将关键位上的数据整个变成None值，后期计算时，做出处理即可!!!
        # date_list={0:[t1,t2..],1:[]...},修改datetime为字典属性
        for voyage_number in date_list.keys():
            k = 0
            datetime_list[voyage_number] = []
            while k < n:  # 这里并不是逐行添加的，注意！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！
                if date_list[voyage_number][k] != 'None':
                    datetime_number = str(date_list[voyage_number][k][:10]) + " " + str(time_list[k])
                    datetime_number_1 = datetime.datetime.strptime(datetime_number, "%Y-%m-%d %H:%M")
                else:
                    datetime_number_1 = 'None'
                datetime_list[voyage_number].append(datetime_number_1)
                k += 1
        # 加入航区的限定！
        # 当时间超过中断时间时跳出
        # 对于判断的时间为空白值的情况，改成手动输入
        if i in c_block:
            k_port_number = p_list.count(k_port)
            #
            if k_port_number == 2:
                break_segment_1 = key_port_address(p_list, k_port)
                # print(strs)
                # print(port_list[break_segment_1[0]] + " " + port_list[break_segment_1[1]])
                # print(port_list[break_segment_2[0]] + " " + port_list[break_segment_2[1]])
                # 遍历航线上的船期
                for voyage_number in datetime_list.keys():
                    impact_time_1 = datetime_list[voyage_number][break_segment_1[0] + 1]
                    impact_time_2 = datetime_list[voyage_number][break_segment_1[1]]
                    # 输入预计到达中断端的时间
                    if impact_time_1 == "None":
                        print(strs)
                        print("row number: " + str(voyage_number))
                        print("column number: " + str(break_segment_1[0] + 1))
                        print("如果超过了，请务必写一个时间！！！")
                        e_date = input("please input date-time as '****-**-** **:**' or 'N'")
                        if e_date != "N":
                            impact_time_1 = datetime.datetime.strptime(e_date, "%Y-%m-%d %H:%M")
                    if impact_time_2 == "None":
                        print(strs)
                        print("row number: " + str(voyage_number))
                        print("column number: " + str(break_segment_1[1]))
                        e_date = input("please input date-time as '****-**-** **:**' or 'N'")
                        if e_date != "N":
                            impact_time_2 = datetime.datetime.strptime(e_date, "%Y-%m-%d %H:%M")
                    # 1在范围之外的跳出、找到一个跳出
                    if impact_time_1 != "None" and impact_time_1 >= disturb_interval[1]:
                        break
                    if impact_time_1 != "None" and disturb_interval[0] <= impact_time_1 <= disturb_interval[1]:
                        delay_time = disturb_interval[1] - impact_time_1
                        k = break_segment_1[0] + 2
                        i_port = p_list[k]
                        t_delay_time = delay_time.days * 24 * 60 * 60 + delay_time.seconds
                        # print(datetime_list[voyage_number][break_segment_1[0] + 2:n])
                        while k < n:
                            if datetime_list[voyage_number][k] != 'None':
                                datetime_list[voyage_number][k] += delay_time
                            k += 1
                        voyage = str(j) + "." + str(voyage_number)
                        delay_voyage.append(voyage)
                        impact_part = [i_port, voyage, t_delay_time]
                        break_impact.append(impact_part)
                        # print(datetime_list[voyage_number][break_segment_1[0] + 2:n])
                        # print("1 " + strs)
                    elif impact_time_2 != "None" and disturb_interval[0] <= impact_time_2 - trans_canal <= \
                            disturb_interval[1]:
                        delay_time = disturb_interval[1] - impact_time_2 + trans_canal
                        k = break_segment_1[1]
                        i_port = p_list[k]
                        t_delay_time = delay_time.days * 24 * 60 * 60 + delay_time.seconds
                        # print(datetime_list[voyage_number][break_segment_1[1]:n])
                        while k < n:
                            if datetime_list[voyage_number][k] != 'None':
                                datetime_list[voyage_number][k] += delay_time
                            k += 1
                        voyage = str(j) + "." + str(voyage_number)
                        delay_voyage.append(voyage)
                        impact_part = [i_port, voyage, t_delay_time]
                        break_impact.append(impact_part)
                        # print(datetime_list[voyage_number][break_segment_1[1]:n])
                        # print("2 " + strs)
            elif k_port_number == 1:  # 检测通过方向
                p = p_list.index(k_port)
                for voyage_number in datetime_list.keys():
                    impact_time_3 = datetime_list[voyage_number][p]
                    if impact_time_3 == "None":
                        print(strs)
                        print("row number: " + str(voyage_number))
                        print("column number: " + str(p))
                        print("如果超过了，请务必写一个时间！！！")
                        e_date = input("please input date-time as '****-**-** **:**' or 'N'")
                        if e_date != "N":
                            impact_time_3 = datetime.datetime.strptime(e_date, "%Y-%m-%d %H:%M")
                    if impact_time_3 != "None" and impact_time_3 >= disturb_interval[1]:
                        break
                    if impact_time_3 != "None" and disturb_interval[0] <= impact_time_3 - trans_canal <= \
                            disturb_interval[1]:
                        delay_time = disturb_interval[1] - impact_time_3 + trans_canal
                        k = p
                        i_port = p_list[k]
                        t_delay_time = delay_time.days * 24 * 60 * 60 + delay_time.seconds
                        # print(datetime_list[voyage_number][p:n])
                        while k < n:
                            if datetime_list[voyage_number][k] != 'None':
                                datetime_list[voyage_number][k] += delay_time
                            k += 1
                        voyage = str(j) + "." + str(voyage_number)
                        delay_voyage.append(voyage)
                        impact_part = [i_port, voyage, t_delay_time]
                        break_impact.append(impact_part)
        # ======================================================================================
        # 利用工作蒲中的数据为算法添加初始的邻接数据
        # 采用p空间的方法存入邻接表数据，即在一条航线上就表示可以直达
        # 航线是环形的，利用错位组合的方法不可避免的会造成不合理的长路径，要避免长路径的出现
        # =============
        port_count = len(p_list)
        # #对于出现列表中数据为空的情况，不添加对应的关键边
        # #理想中的状态是按照航线分别储存航次信息
        # #如果采用列表添加的模式，原本的上层正序、下层倒叙的数据读取方式不再合适
        k = 0
        for voyage_number in datetime_list.keys():
            voyage = strs + "." + str(voyage_number)
            schedule_dict[voyage] = copy.deepcopy(datetime_list[voyage_number])
        datetime_list.clear()
        date_list.clear()
        p_list.clear()
        time_list.clear()


build_file(break_name + "实验\\break impact.json", break_impact)
def extract_text(whole_text):  # 提取船期中航线名的函数
    address = whole_text.index('.')
    results = whole_text[:address]
    return results
def extract_num(whole_text):
    address = whole_text.index('.')
    results = int(whole_text[address + 1:])
    return results
del break_impact
def line_and_datetime_num(aim_dict: dict): # 用于查看航线数目以及船期数目
    line_num = 0
    schedule_num = 0
    for key in aim_dict.keys():
        if "." in key:
            schedule_num += 1
        else:
            line_num += 1
    print("line numbers: " + str(line_num))
    print("schedule numbers: " + str(schedule_num))

# print("schedule")
# line_and_datetime_num(schedule_dict)

def del_schedule(whole_schedule: dict, impact_start: datetime.datetime) -> dict:
    del_list = []
    schedule_part = copy.deepcopy(whole_schedule)
    for key in schedule_part.keys():
        if '.' in key:
            aim_len = len(schedule_part[key]) - 1
            while aim_len >= 0:
                try:
                    if schedule_part[key][aim_len] < impact_start: # 看被检查的船期的最后一个时间是否早于影响开始时间
                        del_list.append(key)
                        break
                    else:
                        break
                except TypeError:
                    aim_len -= 2
    for del_v in del_list:
        if del_v not in delay_voyage:
            try:
                del schedule_part[del_v]
            except KeyError:
                continue
    return schedule_part
remain_schedule = del_schedule(schedule_dict, disturb_interval[0])  # type: dict

def build_od(new_dict: dict, new_load_dict: dict):  # 最新的船期字典，最新的负荷字典，最新的od字典
    new_load_dict.clear()
    new_dict.clear()
    new_dict_part = {}
    for a_voyage in schedule_dict.keys(): # 按照所有的船期生成所有的记录负荷的字典
        if "." in a_voyage:
            aim_schedule = schedule_dict[a_voyage]
            v_name = extract_text(a_voyage)
            ports = schedule_dict[v_name]
            l_1 = 0
            v_length = len(ports)
            while l_1 < v_length:
                l_port = ports[l_1]
                if aim_schedule[l_1] != 'None':  # 没有到达就没有离开
                    l_a = aim_schedule[l_1]
                    l_d = aim_schedule[l_1 + 1]
                    # 港口到达或者离开时间，仅仅通过时间区分，不唯一！需要更改
                    if l_port in new_load_dict.keys():
                        new_load_dict[l_port][(a_voyage, l_a)] = [0, l_d - l_a, -1]
                        new_load_dict[l_port][(a_voyage, l_d)] = [0, l_d - l_a, 1]
                    else:
                        new_load_dict[l_port] = {}
                        new_load_dict[l_port][(a_voyage, l_a)] = [0, l_d - l_a, -1]
                        new_load_dict[l_port][(a_voyage, l_d)] = [0, l_d - l_a, 1]
                    # 对应的数据结构是，load_dict={港口名：{到达时间：[计数，卸货时间区间,航次]，离开时间：[计数，装货时间区间,航次】，。。。}}
                    # 更换一下初始的数据
                l_1 += 2

    for b_voyage in remain_schedule.keys():
        p_1 = 0
        if "." in b_voyage:
            aim_schedule = remain_schedule[b_voyage]
            v_name = extract_text(b_voyage)
            ports = remain_schedule[v_name]
            v_length = len(ports)
            while p_1 < v_length - 2:  # 点坐标2n，列表长度2n+1
                if aim_schedule[p_1 + 1] != 'None':
                    s_point = ports[p_1]
                    p_2 = v_length - 2
                    while p_2 > p_1:  # 船期中od信息只能是向后的,所以在起始点后一位结束
                        if aim_schedule[p_2] != 'None':
                            e_point = ports[p_2]
                            if s_point != e_point:  # 删除掉本地到本地的情况
                                d_point = p_1 + 1
                                a_point = p_2
                                d_time = aim_schedule[d_point]
                                a_time = aim_schedule[a_point]
                                # 改变line_dict
                                new_dict_part[s_point, e_point] = [s_point, e_point, d_time, a_time, b_voyage]
                                # new_load_dict[s_point][(d_time, 1)][0] += 1
                                # new_load_dict[e_point][(a_time, -1)][0] += 1
                                # 原本是航线字典={o:[]}
                                # 将当前航线中的od信息导入:whole_od_list{(O,D):[[[1]],[[2]],[[3]]...]}，(O,D):[[1],[2]...].....}
                            p_2 -= 2
                        else:
                            p_2 -= 2
                    p_1 += 2
                else:
                    p_1 += 2
            for o in new_dict_part.keys():
                if o in new_dict.keys():
                    new_dict[o].append(copy.deepcopy([new_dict_part[o]]))
                else:
                    new_dict[o] = []
                    new_dict[o].append(copy.deepcopy([new_dict_part[o]]))
            new_dict_part.clear()

# line_and_datetime_num(schedule_dict)
def build_schedule(json_name: str, s_dict: dict):  # 将保存船期的字典存为json文件
    temp_dict = {}
    for key in s_dict.keys():
        temp_dict[key] = []
        if "." in key:
            for v_time in s_dict[key]:
                if v_time != 'None':
                    temp_dict[key].append(datetime.datetime.strftime(v_time, "%Y-%m-%d %H:%M"))
                else:
                    temp_dict[key].append(v_time)
        else:
            temp_dict[key] = s_dict[key]
    build_file(json_name, temp_dict)
# print("or")
# line_and_datetime_num(schedule_dict)
# print("remain")
# line_and_datetime_num(remain_schedule)

# 删除无效的船期

def load_counter(od_path: list, count_can: dict):
    for part in od_path:
        p_1 = part[0]
        # [start_point,end_point,depart_time_1, arrive_time_1,voyage]
        p_1_leave = part[2]
        p_2 = part[1]
        p_2_arrive = part[3]
        l_m = part[-1]
        count_can[p_1][(l_m, p_1_leave)][0] += 1
        count_can[p_2][(l_m, p_2_arrive)][0] += 1
# 每一次更新路径生成一个记录路径变化的字典，在更新路径的函数中嵌入对以往缺失路径的修复？
def detect_path(aim_dict: dict) -> list:# 检测初始路径存续情况
    update_list = []
    if aim_dict:
        for key in aim_dict.keys():
            c_num = 0
            for path_state in aim_dict[key]:
                c_num += path_state[0]
                c_num += path_state[1]
            if c_num < 0: # 有断路
                update_list.append(key)
    return update_list
change_times = 0
path_change = {}
def detect_num(change_dict):
    change_num = 0
    for key in change_dict:
        change_num += change_dict[key][0]
        change_num += change_dict[key][1]
    return change_num
# 记录负载情况
# 更新一次记录一次json文件
def change_way(change_name): # 更改路径+路径检查？先更新所有变换的路径单元、找到所有中断的路径、全局更新后续的路径
    decrease_num = 0
    for key in or_whole_od.keys():
        change_list = []# [增加路径，减少路径，变换所在路径的id，中断位于的港口，变换所在的次数]
        p_address = 0
        for path in or_whole_od[key]:
            for part in path:
                if part[-1] in delay_voyage:
                    change_od = (part[0],part[1])
                    for base_od in new_od_dict[change_od]:
                        if base_od[0][-1] == part[-1]:
                            part[2] = copy.deepcopy(base_od[0][2])
                            part[3] = copy.deepcopy(base_od[0][3])
            p_index = 0
            while p_index < len(path) - 1:
                if path[p_index][3] + handing_time >= path[p_index + 1][2]:
                    # 记录缺少的路径
                    if change_list:
                        change_list[1] -= 1
                        change_list[2].append(p_address)
                        change_list[3].append([path[p_index][1],path[p_index][-1]])
                        for a_part in path[p_index + 1:]:
                            change_list[5].append([a_part[0],a_part[1],a_part[-1]])
                        decrease_num += 1
                    else:
                        change_list = [0, 0, [],[],change_times,[],[]]#[增加路径，减少路径，[影响的路径单元位置],原始的后半段路径，新的后半段路径]
                        change_list[1] -= 1
                        change_list[2].append(p_address)
                        change_list[3].append([path[p_index][1],path[p_index][-1]])
                        for a_part in path[p_index + 1:]:
                            change_list[5].append([a_part[0], a_part[1], a_part[-1]])
                        decrease_num += 1
                    del path[p_index + 1:]
                    break#while循环的是路径单元，最早出现断裂，就要跳出，并找下一个路径中的中断情况
                p_index += 1
            p_address += 1
        if change_list:
            path_change[key] = copy.deepcopy(change_list)
            change_list.clear()
    # 出现影响的路径，进行路径更新（"a","b"):[0,0,[],change times]
    find_times = 0
    while decrease_num > 0:
        print(find_times) # 搜索路径的次数
        check_m = detect_num(path_change)
        for key in path_change.keys():
            cut_list = []
            if path_change[key][2]:
                for diff_index in path_change[key][2]:
                    arrive = or_whole_od[key][diff_index][-1][3]
                    after = (or_whole_od[key][diff_index][-1][1],key[1])
                    t_list = []
                    for after_path in or_whole_od[after]:
                        if after_path[0][2] > arrive + handing_time:
                            if t_list:
                                if after_path[-1][3] < t_list[-1][3]:
                                    t_list = after_path
                            else:
                                t_list = after_path
                    if t_list:
                        for l_part in t_list:
                            or_whole_od[key][diff_index].append(l_part)
                            path_change[key][6].append([l_part[0],l_part[1],l_part[-1]])
                        path_change[key][0] += 1
                        decrease_num -= 1
                        cut_list.append(diff_index)
                if cut_list:
                    for cut_num in cut_list:
                        path_change[key][2].remove(cut_num)
        find_times += 1
        check_n = detect_num(path_change)
        if check_m == check_n:
            break
    change_sign = 0
    path_change_2 = {}
    for key in path_change:
        path_change_2[change_sign] = list(key) + path_change[key]
        change_sign += 1
    build_file(change_name,path_change_2)
    path_change.clear()
    del path_change_2
    for key in or_whole_od.keys():
        for path in or_whole_od[key]:
            load_counter(path,load_dict)

def build_interval_can(years: int):  # 按照某年生成一个区间树容器
    month_num = list(range(1, 13))  # 生成月份列表
    clock_list = list(range(0, 48))  # 生成时间区间列表
    times = dict.fromkeys(month_num, {})
    for m in times.keys():
        monthrange = calendar.monthrange(years, m)  # 按照某年的月份取一个月的天数
        d = monthrange[-1]  # 按照2020年的月份取一个月的天数
        day_list1 = list(range(1, d + 1))  # 按照天数生成列表
        times[m] = dict.fromkeys(day_list1, {})
    for m in times.keys():
        for d in times[m].keys():
            times[m][d] = dict.fromkeys(clock_list)
            for t in times[m][d].keys():
                times[m][d][t] = [0, 0, [], []]  # 初始化为0
            # times[j][k] = dict.fromkeys(time_list,[0,0])
            # 列表中0位置上存放卸货负荷，1位上存放装货负荷,2位上存放到港船期，3位置上存放离港船期，
    return times
# '''生成区间树容器'''
# #按照港口划分区间树，需要调整！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！
def count_load(load_dicts: dict):  # time_unit是累加的时间单元，默认为半小时
    # 负荷字典、港口初始负载字典、港口区间树字典、过载港口字典
    # 生成港口区间树负荷统计字典的函数
    port_interval_tree = {}
    for l_port in load_dicts.keys():
        times_can = build_interval_can(2020)
        for t_label in load_dicts[l_port].keys():
            # 计算时间差有多少个半小时
            time_diff = load_dicts[l_port][t_label][1]
            half_num = time_diff.days * 48 + time_diff.seconds / 1800  # 计算有几个半小时
            # 计算实际负荷
            true_load = load_dicts[l_port][t_label][0] / half_num
            t_count = t_label[1]
            if load_dicts[l_port][t_label][2] < 0:
                while t_count < t_label[1] + time_diff:  # 到达卸货负荷，一直到船离开
                    # 按照访问时间，获取月份、天
                    month_num = t_count.month
                    date_num = t_count.day
                    hour_num = t_count.hour * 2 + t_count.minute / 30
                    # 这个取大小，比较对象也有问题
                    times_can[month_num][date_num][hour_num][0] -= true_load
                    times_can[month_num][date_num][hour_num][2].append(t_label)
                    t_count += datetime.timedelta(minutes=30)
                    # 因为load_dict中的时间戳不是排序的，所以利用函数区间函数进行统计的时候，取最大可能是局部的最大
            else:
                t_count -= datetime.timedelta(minutes=30)
                while t_count >= t_label[1] - time_diff:
                    month_num = t_count.month
                    date_num = t_count.day
                    hour_num = t_count.hour * 2 + t_count.minute / 30
                    times_can[month_num][date_num][hour_num][1] += true_load
                    times_can[month_num][date_num][hour_num][3].append(t_label)
                    t_count -= datetime.timedelta(minutes=30)
        port_interval_tree[l_port] = copy.deepcopy(times_can)
        del times_can
    return port_interval_tree
    # 生成港口节点每个时间区间的负荷累计量
def early_overload(port_interval_tree: dict, original_load_dict: dict, ratio: float) -> list:  # 输出储存过载时段的信息，输出嵌套列表
    # （记录负载的字典,港口初始负荷，因素，过载路径，上一次过载字典(start_time,end_time,port)）
    # 两个节点同时发生过载，过载之间是否相互影响？
    early_list = []
    for p_name in port_interval_tree.keys():
        new_load_dict = copy.deepcopy(port_interval_tree[p_name])
        port_max_load = original_load_dict[p_name][1] - original_load_dict[p_name][0]
        for m in new_load_dict:
            for d in new_load_dict[m]:
                for t in new_load_dict[m][d]:
                    aim_port_load = new_load_dict[m][d][t][1] - new_load_dict[m][d][t][0]
                    if aim_port_load > port_max_load * ratio:
                        if t % 2 == 0:
                            minute = "00"
                        else:
                            minute = "30"
                        clock1 = "2020-" + str(m) + "-" + str(d) + " " + str(t // 2) + ":" + minute
                        clock2 = datetime.datetime.strptime(clock1, "%Y-%m-%d %H:%M")
                        # 判断时间、港口的情况，如果不是同一个港口
                        # if last_overload:
                        #     if p_name == last_overload[2]:
                        #         if clock2 > last_overload[1]:
                        #             note_list = [clock2, p_name, new_load_dict[m][d][t][0], new_load_dict[m][d][t][1],
                        #                          new_load_dict[m][d][t][2], new_load_dict[m][d][t][3]]
                        #         else:
                        #             break
                        #     else:
                        #         note_list = [clock2, p_name, new_load_dict[m][d][t][0], new_load_dict[m][d][t][1],
                        #                      new_load_dict[m][d][t][2], new_load_dict[m][d][t][3]]
                        #     # [时刻，港口名，卸货负荷，装货负荷，到达时间，离开时间]
                        #
                        #     if early_list:
                        #         early_time = early_list[0][0]
                        #         note_time = note_list[0]
                        #         if note_time < early_time:
                        #             early_list.clear()
                        #             early_list.append(copy.deepcopy(note_list))
                        #         else:
                        #             if note_list[1] == early_list[0][1]:
                        #                 if note_list[2] == early_list[0][2]:
                        #                     if note_list[3] == early_list[0][3]:
                        #                         early_list.append(copy.deepcopy(note_list))
                        #             else:
                        #                 break
                        #     else:
                        #         early_list.append(copy.deepcopy(note_list))
                        # else:
                        note_list = [clock2, p_name, new_load_dict[m][d][t][0], new_load_dict[m][d][t][1],
                                     new_load_dict[m][d][t][2], new_load_dict[m][d][t][3],aim_port_load,port_max_load]
                        if early_list:
                            early_time = early_list[0][0]
                            note_time = note_list[0]
                            if note_time < early_time:
                                early_list.clear()
                                early_list.append(copy.deepcopy(note_list))
                            else:
                                if note_list[1] == early_list[0][1]:
                                    if note_list[2] == early_list[0][2]:
                                        if note_list[3] == early_list[0][3]:
                                            early_list.append(copy.deepcopy(note_list))
                                else:
                                    break
                        else:
                            early_list.append(copy.deepcopy(note_list))
        del new_load_dict
    if early_list:
        # last_overload.clear()
        # gc.collect()
        # last_overload.append(copy.deepcopy(early_list[0][0]))
        # last_overload.append(copy.deepcopy(early_list[-1][0]))
        # last_overload.append(copy.deepcopy(early_list[0][1]))

        overload_s = datetime.datetime.strftime(early_list[0][0], "%Y-%m-%d %H:%M")
        overload_e = datetime.datetime.strftime(early_list[-1][0] + datetime.timedelta(minutes=30), "%Y-%m-%d %H:%M")
        # if p_name in overload_way.keys():
        early_port = early_list[0][1]

        overload_way[change_times] = [early_port,overload_s, overload_e, [], [],early_list[0][6],early_list[0][7]] # 3上的方括号是延长离港的船期，后一个是等待靠泊的船期
    return early_list  # 输出的是一个港口过载的情况此时造成过载的到港船舶情况是一致的
# 前方需要将直达路径与中转路径分开！
# 对于港口出现过载的情况分为两种1、只有一个航期 2、有多个航期
# 有一个航期的我们将离港时间延长
# 定义一个确认过载的船期的情况的函数
# 中断的时间应该是过载的终点
def check_voyage(early_list_part: list, latest_load):
    # 找到的最早出现过载的情况，记录了时间、港口名，卸货负荷，装货负荷，到达时间，离开时间
    overload_dict = {}
    p_name = early_list_part[1]
    # [clock1, p_name, new_load_dict[m][d][t][0],new_load_dict[m][d][t][1],new_load_dict[m][d][t][2], new_load_dict[m][d][t][3]]
    if early_list_part[4]:
        for table in early_list_part[4]:
            c_voyage = table[0]
            unload = latest_load[p_name][table][0]
            overload_dict[c_voyage] = [table[1], 0, unload, 0, c_voyage]
    if early_list_part[5]:
        for table in early_list_part[5]:
            c_voyage = table[0]
            load = latest_load[p_name][table][0]
            if c_voyage in overload_dict.keys():
                overload_dict[c_voyage][1] = table[1]
                overload_dict[c_voyage][3] = load
            else:
                overload_dict[c_voyage] = [0, table[1], 0, load, c_voyage]
                a_time = table[1] - latest_load[p_name][table][1]
                early_list_part[c_voyage][0] = a_time
    overload_list_cache = list(overload_dict.values())
    overload_list_cache.sort(key=lambda x: x[0], reverse=False)
    return overload_list_cache  # 输出一个嵌套列表，包含过载期间所有的船期，以及船期对应的抵离港时间、装卸负载以及船期名
# over_load:过载的负荷，每半小时多少负荷，overload_time：多少个半小时

def one_voyage_change(overload_list_cache, over_load, overload_time, p_or_load):
    print("one voyage change!")
    aim_voyage = overload_list_cache[0][-1]
    d_time = overload_list_cache[0][1]
    voyage_l = len(schedule_dict[aim_voyage])
    p_index = schedule_dict[aim_voyage].index(d_time)
    p_time = (math.ceil(((over_load * overload_time) / p_or_load - overload_time))) * half_hour
    t_p_time = p_time.days * 24 * 60 * 60 + p_time.seconds
    overload_way[change_times][3].append([aim_voyage, t_p_time])  # 1表示离港时间延长
    while p_index < voyage_l:
        if schedule_dict[aim_voyage][p_index] != 'None':
            schedule_dict[aim_voyage][p_index] += p_time
        p_index += 1
    delay_voyage.append(aim_voyage)

def multiple_voyage_change(overload_list_cache, overload_time, overload_end, p_or_load):
    # （过载的时刻的船期从早到晚排序，最新的船期，过载时间记录情况，该港口的初始负荷）过载以后的节点不予理会
    # [[船期，（0/1）等待/延迟，延后的时间，到达时间/离开时间]]
    print("multiple voyage change!")
    check_load = p_or_load
    count_load_point = 0
    while check_load > 0:
        aim_voyage = overload_list_cache[count_load_point]
        # 估计该船期的负荷
        diff_half = (aim_voyage[1] - aim_voyage[0]).days * 48 + (aim_voyage[1] - aim_voyage[0]).seconds / 1800# 计算有多少个半小时
        t_l = (aim_voyage[2] + aim_voyage[3]) / diff_half
        check_load -= t_l
        count_load_point += 1
    # 找到临界值的船期，船期在港时间延长一部分！！
    t_deal = t_l + check_load
    # 静态负荷处理准则，一旦船舶靠泊，其分配的处理能力不变
    d_time = math.ceil((aim_voyage[2] + aim_voyage[3]) / t_deal) * half_hour + aim_voyage[0]
    w_time = overload_end - aim_voyage[0] + aim_voyage[1]
    # 对比是等待快还是用仅有的资源做的快
    if d_time <= w_time:
        p_t = d_time - aim_voyage[1]  # 该船期在港延长的时间
        t_p_t = p_t.days * 24 * 60 * 60 + p_t.seconds# 需要记录过载影响，将时间戳转化为秒
        l_time = aim_voyage[1]
        prolong_voyage = aim_voyage[-1]
        t_index = schedule_dict[prolong_voyage].index(l_time)
        delay_voyage.append(prolong_voyage)
        overload_way[change_times][3].append([prolong_voyage, t_p_t])
        while t_index < len(schedule_dict[prolong_voyage]):
            if schedule_dict[prolong_voyage][t_index] != 'None':
                schedule_dict[prolong_voyage][t_index] += p_t
            t_index += 1
    else:
        count_load_point -= 1
    # overload_list_cache[count_load_point - 1][1] += p_t #同步更新临时的过载记录列表[[a_time,d_time,unload,load,voyage_number]]

    while count_load_point < len(overload_list_cache):
        remain_list = overload_list_cache[count_load_point]
        w_voyage = remain_list[-1]
        delay_voyage.append(w_voyage)
        w_index = schedule_dict[w_voyage].index(remain_list[0])  # 找到等待时间的索引
        overload_way[change_times][4].append([w_voyage, overload_time * 1800])
        while w_index < len(schedule_dict[w_voyage]):
            if schedule_dict[w_voyage][w_index] != 'None':
                schedule_dict[w_voyage][w_index] += overload_time * half_hour
            w_index += 1
        count_load_point += 1
    # 从索引位置开始延长等待时间
    # o_e表示多个过载船期中，最早到达的船期：[a_time, d_time, unload, load, overload_voyage]
    # 先入先出准则？过载的部分延长

build_od(new_od_dict,load_dict)


# way_num = 0
# for key in new_od_dict.keys():
#     way_num += len(new_od_dict[key])
# print(way_num)
# sys.exit()
# build_path(new_od_dict)

def build_whole_od(json_name, whole_od):
    temp_dict = {}
    kk = 0
    for o_d in whole_od.keys():
        temp_dict[kk] = copy.deepcopy(whole_od[o_d])
        for od_path in temp_dict[kk]:
            for path_part in od_path:
                try:
                    path_part[2] = datetime.datetime.strftime(path_part[2], "%Y-%m-%d %H:%M")
                except TypeError:
                    continue
                try:
                    path_part[3] = datetime.datetime.strftime(path_part[3], "%Y-%m-%d %H:%M")
                except TypeError:
                    continue
        kk += 1
    build_file(json_name, temp_dict)

# build_whole_od(break_name + "实验\\after break whole od.json", new_od_dict)

or_json = load_file("whole od path (key as number).json")
# 将记录初始路径信息的字典恢复成可用的字典数据
or_whole_od = recover_whole_od(or_json)
# 更新路径，将上一次的最短路径中改动的路径添加到新的路径集中
del or_json
change_json = break_name + "实验\\after " + str(change_times) + " change way.json"

change_way(change_json)

#change_way(or_whole_od, new_od_dict, disturb_interval[0])

print("finish after break change way!")

port_interval = count_load(load_dict)

or_load_dict = load_file("port max load.json")

factor = 1

overload_list = early_overload(port_interval, or_load_dict, factor)

print("finish overload find after break!")
overload_voyage_dict = {}
overload_times = 1
while len(overload_list) > 0:
    overload_part = overload_list[0]

    overload_voyage = check_voyage(overload_part, load_dict)

    print(overload_voyage)
    voyage_list = []
    for message in overload_voyage:
        d_time = (message[1] - message[0]).days*48 + (message[1] - message[0]).seconds/1800
        average_load = (message[2] + message[3])/d_time
        voyage_list.append([message[-1],average_load])
    overload_voyage_dict[overload_times] = copy.deepcopy(voyage_list)

    voyage_list.clear()
    overload_times += 1

    port_name = overload_part[1]
    overload_period = len(overload_list)
    overload_now = overload_part[3] - overload_part[2]
    or_load = or_load_dict[port_name][1] - or_load_dict[port_name][0]

    delay_voyage.clear()
    over_start_time = overload_list[0][0]
    over_end_time = overload_list[-1][0]

    if len(overload_voyage) == 1:
        one_voyage_change(overload_voyage, overload_now, overload_period, or_load)
    else:
        multiple_voyage_change(overload_voyage, overload_period, over_end_time, or_load)

    print("finish schedule change!")

    # build_schedule(break_name + "实验\\第" + str(change_times) + "次船期变化后船期表.json", schedule_dict)

    del remain_schedule

    remain_schedule = del_schedule(schedule_dict, over_start_time)  # type: dict
    build_od(new_od_dict,load_dict)  # 清除new_od_dict中的内容，清除load_dict中的内容

    change_times += 1
    change_json = break_name + "实验\\after " + str(change_times) + " change way.json"
    change_way(change_json)

    port_interval.clear()

    port_interval = count_load(load_dict)

    overload_list = early_overload(port_interval, or_load_dict, factor)

    print("finish overload find after schedule change!")

build_file(break_name + "实验\\overload way.json", overload_way)


overload_voyage_file = break_name + "实验\\overload voyage.json"
build_file(overload_voyage_file,overload_voyage_dict)

def change_dict(aim_dict):
    new_dict = {}
    k = 0
    for od in aim_dict.keys():
        new_dict[k] = aim_dict[od]
        k += 1
    return new_dict


def trans_whole_od(aim_dict):

    for o_d in aim_dict.keys():
        for od_path in aim_dict[o_d]:
            for path_part in od_path:
                try:
                    path_part[2] = datetime.datetime.strftime(path_part[2],"%Y-%m-%d %H:%M")
                except TypeError:
                    continue
                try:
                    path_part[3] = datetime.datetime.strftime(path_part[3],"%Y-%m-%d %H:%M")
                except TypeError:
                    continue
    return aim_dict

new_whole_od = trans_whole_od(or_whole_od)
new_key_path = change_dict(new_whole_od)

aim_json = break_name + "实验\\whole od path (key as number).json"
build_file(aim_json,new_key_path)
end_time = time.time()
print("use " + str(end_time - start_time) + "seconds")
