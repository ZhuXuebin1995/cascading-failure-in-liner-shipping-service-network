#在计算最短路的基础上，统计各个节点区间的访问次数
#coding:utf-8
import openpyxl #导入工作蒲模块
from openpyxl.utils import get_column_letter #导入模块中读取最大列的函数
from os import listdir
import os
import datetime
#设置的变量务必不可以跟模块、函数等重名！！！
import time
import copy
import pandas as pd
import calendar
import json
start_time = time.time()
#===================================================================================================================
#为算法加载初步的数据
#将文件夹中的所有工作簿的工作表以邻接、直达的路径时间导入短期边中，边包含所属航线信息
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

path_list = []
open_file_path = r'E:\大连海事大学\博士\数据\集装箱船公司船期数据\中远海集运（按照区域航线划分）\标准化数据'
file_name = listdir(open_file_path)
line_dict = {}#按照航线名创建嵌套中的键值对，此时的结构是：航线字典 = {od:[],od:[]}
for i in file_name:
    com_path = os.path.join(open_file_path,i)
    path_list.append(com_path)
#print(path_list)
#path_list中已经有所有的工作蒲地址
port_list = [] #临时表
l_num = 0
line_dict= []
for i in path_list:#遍历每一个工作蒲
    wb = openpyxl.load_workbook(i)
    sheet_names = wb.sheetnames # 获取工作蒲中所有表格
    #从表格中提取港口名、时刻、日期存入三个临时列表（重置）

    for j in sheet_names:
        sheet = wb[j]
        max_column = sheet.max_column
        max_row = sheet.max_row #获取最大行数
        column = get_column_letter(max_column)
        row_1 = sheet["I3":"%s3" % column]

        for row1_cells in row_1:
            for cell in row1_cells:
                if str(cell.value) != 'None':
                    port_list.append(str(cell.value))#读取出每个工作表的港口信息
        line_dict.append(copy.deepcopy(port_list))
        port_list.clear()
    l_num += 1
port_route_num = {}
temp_list = []
port_go_dict = {}
check_list = []
for line in line_dict:
    line_length = len(line)
    k = 0
    while k < line_length - 1:
        temp_port = line[k]
        if temp_port not in port_go_dict.keys():
            port_go_dict[temp_port] = []
        if temp_port in check_list:
            k += 1
            continue
        else:
            check_list.append(temp_port)
            temp_line = line[k+1:]
            for p in temp_line:
                if p not in port_go_dict[temp_port]:
                    if p != temp_port:
                        port_go_dict[temp_port].append(p)
            k += 1
    check_list.clear()
    for port in line:
        if port not in temp_list:
            if port not in port_route_num.keys():
                port_route_num[port] = 1
            else:
                port_route_num[port] += 1
            temp_list.append(port)
        
    temp_list.clear()

wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1).value = "port"
ws1.cell(row=1, column=2).value = "route number"
ws1.cell(row=1, column=3).value = "go_port"
r = 2  # type: int
for port in port_route_num:
    ws1.cell(row=r, column=1).value = port
    ws1.cell(row=r, column=2).value = port_route_num[port]
    ws1.cell(row=r, column=3).value = len(port_go_dict[port])
    r += 1


wb.save(filename="ports' line number.xlsx")
