from openpyxl.utils import get_column_letter
from openpyxl import Workbook
'''
{"0": ["YANTIAN", "ENSENADA", 1, -1, [], ["KAOHSIUNG","SHANGHAI"], 0],
'''
import json
def load_file(address):
    file_address = address
    with open(file_address) as add_json:
        new_dict = json.load(add_json)
    return new_dict
break_time = "中断8天实验"
times = 0
change_news = {"cut_num": 0, "add_num": 0}
diff_num = 0

while True:
    change_way_file = break_time + "\\after " + str(times) + " change way.json"
    try:
        change_way = load_file(change_way_file)
        if change_way_file:
            for key in change_way:
                change_news["cut_num"] += change_way[key][3]
                change_news["add_num"] += change_way[key][2]
                # for impact_port in change_way[key][5]:
                #     if impact_port in change_news.keys():
                #         change_news[impact_port] += 1
                #     else:
                #         change_news[impact_port] = 1
                # if change_way[key][3] + change_way[key][2] != 0:
                #     a = "change_diff" + str(diff_num)
                #     change_news[a] = str([change_way[key][0],change_way[key][1],change_way[key][2],change_way[key][3],change_way[key][5]])
                #     diff_num += 1
        times += 1
    except FileNotFoundError:
        break
#
wb = Workbook()
ws1 = wb.active
number = 1
aim_file_name = break_time + "\\change news.xlsx"
for item in change_news.keys():
    ws1.cell(row=number, column=1).value = item
    ws1.cell(row=number, column=2).value = change_news[item]
    number += 1
wb.save(filename=aim_file_name)

