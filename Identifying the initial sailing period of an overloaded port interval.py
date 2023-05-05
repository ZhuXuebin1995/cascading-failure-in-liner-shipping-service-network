#载入过载情况
#载入船期数据
#遍历对照，输出船期情况

import json
with open('中断1天实验\\overload way.json') as json_file:
    overload_message = json.load(json_file)
print(overload_message)

# 港口、船期在港时间、船期数目
# 时间区间进行船期统计，一个时间段出现船期+1
# 船期统计次数
from typing import List, Any

break_time = "中断1天实验"
'''break_name + "实验\\break impact.json",'''

from os import listdir
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

open_file_path = r'E:\大连海事大学\博士\数据\集装箱船公司船期数据\中远海集运（按照区域航线划分）\标准化数据'
file_name = listdir(open_file_path)

p_list = []
time_list = []
date_list = {}
port_interval = {}
port_voyage_count = {}
datetime_list = {}
import datetime
import calendar


def build_interval_can(years: int):  # 按照某年生成一个区间树容器
    month_num = list(range(1, 13))  # type: List[Any] # 生成月份列表
    clock_list = list(range(0, 48))  # 生成时间区间列表
    times = dict.fromkeys(month_num, {})
    for m in times.keys():
        monthrange = calendar.monthrange(years, m)  # 按照某年的月份取一个月的天数
        d = monthrange[-1]  # 按照2020年的月份取一个月的天数
        day_list1 = list(range(1, d + 1))  # 按照天数生成列表
        times[m] = dict.fromkeys(day_list1, {})
    for m in times.keys():
        for d in times[m].keys():
            times[m][d] = dict.fromkeys(clock_list)
            for t in times[m][d].keys():
                times[m][d][t] = []  # 初始化为0
            # times[j][k] = dict.fromkeys(time_list,[0,0])
            # 列表中0位置上存放卸货负荷，1位上存放装货负荷,2位上存放到港船期，3位置上存放离港船期，
    return times


for i in file_name:
    com_path = os.path.join(open_file_path, i)
    wb = openpyxl.load_workbook(com_path)
    sheet_names = wb.sheetnames  # 获取工作蒲中所有表格
    for j in sheet_names:  # 从表格中提取港口名、时刻、日期存入三个临时列表（重置）
        strs = str(j)  # 按照航线名称创建字典键
        sheet = wb[j]
        max_column = sheet.max_column
        max_row = sheet.max_row  # 获取最大行数
        column = get_column_letter(max_column)
        row_1 = sheet["I3":"%s3" % column]
        for row1_cells in row_1:
            for cell in row1_cells:
                p_list.append(str(cell.value))  # 读取出每个工作表的港口信息
        row_2 = sheet["I5":"%s5" % column]
        for row2_cells in row_2:
            for cell in row2_cells:
                time_list.append(str(cell.value))
        for row_number in range(6, max_row + 1):  # 遍历行，对row3做修正==修正！！！！成立一个关日期行的字典？后期添加航线字典也使用字典值的遍历！！！！！
            first_column = str("I" + str(row_number))
            last_column = str("%s" + str(row_number))
            row_3 = sheet[first_column:last_column % column]  # row_3 = sheet["I6":"%s6" % column]
            voyage_number = len(date_list)
            date_list[voyage_number] = []
            for row3_cells in row_3:
                for cell in row3_cells:
                    date_list[voyage_number].append(str(cell.value))
        for port in p_list:  # 生成真实的港口列表
            if port not in port_interval:
                if port != 'None':
                    port_interval[port] = build_interval_can(2020)


                    # 计划设计{}中的数据为time：[次数，装卸货时间]当前是到达时间，则是卸货，前时间-后时间符号为负的；当前是离开时间，则是装货，后时间-前时间，符号为正的
                    # 不确定这样的赋值方式是否适用！！！！！！！！！！
        n = len(p_list)
        # print(str(n))
        # 在检查结果的时候，出现"None 13:00"的字样，原因是有航次在这个日期没有船舶靠泊的安排，但是不能将其从列表中删除，否则出现列表串位的情况
        # 因此在发现有None的时候，将关键位上的数据整个变成None值，后期计算时，做出处理即可!!!
        # date_list={0:[t1,t2..],1:[]...},修改datetime为字典属性
        for voyage_number in date_list.keys():
            k = 0
            datetime_list[voyage_number] = []
            while k < n:  # 这里并不是逐行添加的，注意！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！
                if date_list[voyage_number][k] != 'None':
                    datetime_number = str(date_list[voyage_number][k][:10]) + " " + str(time_list[k])
                    datetime_number_1 = datetime.datetime.strptime(datetime_number, "%Y-%m-%d %H:%M")
                else:
                    datetime_number_1 = 'None'
                datetime_list[voyage_number].append(datetime_number_1)
                k += 1
        for voyage_number in datetime_list.keys():
            k = 0
            voyage_name = j + '.' + str(voyage_number)
            while k < n:
                port = p_list[k]
                if datetime_list[voyage_number][k] != 'None':
                    v_s = datetime_list[voyage_number][k]
                    v_e = datetime_list[voyage_number][k + 1]
                    while v_s < v_e:
                        month_num = v_s.month
                        date_num = v_s.day
                        hour_num = v_s.hour * 2 + v_s.minute / 30
                        port_interval[port][month_num][date_num][hour_num].append(voyage_name)
                        v_s += datetime.timedelta(minutes=30)
                k += 2
        datetime_list.clear()
        date_list.clear()
        p_list.clear()
        time_list.clear()




for i in range(1, 9):
    original_voyages = []
    break_name = "中断" + str(i) + "天实验"
    overload_file_address = break_name + "\\overload or load.json"
    with open(overload_file_address) as file_json:
        overload_file = json.load(file_json)
    for key in overload_file.keys():
        port_name = overload_file[key][0]
        time1 = overload_file[key][1]
        time1_1 = datetime.datetime.strptime(time1,'%Y-%m-%d %H:%M')
        month = time1_1.month
        day = time1_1.day
        hour_minute = time1_1.hour*2 + time1_1.minute/30
        original_voyages.append(port_interval[port_name][month][day][hour_minute])

    json_build = json.dumps(original_voyages)
    original_voyage_address = break_name + "\\overload port original voyage.json"
    with open(original_voyage_address,'w') as json_can:
        json_can.write(json_build)