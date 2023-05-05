# "0": ["QINGDAO", "2020-05-19 20:00", "2020-05-19 22:30", [], [["AWE1.19", 10800]]],
break_time = "中断3天实验"
from openpyxl import Workbook
import datetime
import json
from openpyxl.utils import get_column_letter
def load_file(address):
    file_address = address
    with open(file_address) as add_json:
        new_dict = json.load(add_json)
    return new_dict
overload_way_address = break_time + "\\overload way.json"
overload_way = load_file(overload_way_address)
wb = Workbook()
ws1 = wb.active
overload_file = break_time + "\\过载情况.xlsx"
wb = Workbook()
ws1 = wb.active

ws1.cell(row=1,column=1).value = "port"
ws1.cell(row=1,column=2).value = "overload start time"
ws1.cell(row=1,column=3).value = "overload end time"
ws1.cell(row=1,column=4).value = "state"
ws1.cell(row=1,column=5).value = "voyage"
ws1.cell(row=1,column=6).value = "add time"
ws1.cell(row=1,column=7).value = "delta overload time"
r = 2
for key in overload_way.keys():
    ws1.cell(row=r, column=1).value = overload_way[key][0]
    ws1.cell(row=r, column=2).value = overload_way[key][1]
    ws1.cell(row=r, column=3).value = overload_way[key][2]
    t1 = datetime.datetime.strptime(overload_way[key][1],"%Y-%m-%d %H:%M")
    t2 = datetime.datetime.strptime(overload_way[key][2],"%Y-%m-%d %H:%M")
    delta_time = (t2-t1).days*24+(t2-t1).seconds/3600
    ws1.cell(row=r, column=7).value = delta_time

    if overload_way[key][3]:
        ws1.cell(row=r, column=4).value = "延长离开"
        ws1.cell(row=r, column=5).value = overload_way[key][3][0][0]
        ws1.cell(row=r, column=6).value = overload_way[key][3][0][1]
        r += 1
        if overload_way[key][4]:
            ws1.cell(row=r, column=1).value = overload_way[key][0]
            ws1.cell(row=r, column=2).value = overload_way[key][1]
            ws1.cell(row=r, column=3).value = overload_way[key][2]
            ws1.cell(row=r, column=4).value = "推迟到达"
            ws1.cell(row=r, column=5).value = overload_way[key][4][0][0]
            ws1.cell(row=r, column=6).value = overload_way[key][4][0][1]
            t1 = datetime.datetime.strptime(overload_way[key][1], "%Y-%m-%d %H:%M")
            t2 = datetime.datetime.strptime(overload_way[key][2], "%Y-%m-%d %H:%M")
            delta_time = (t2 - t1).days * 24 + (t2 - t1).seconds / 3600
            ws1.cell(row=r, column=7).value = delta_time
            r += 1
            continue
    if overload_way[key][4]:
        ws1.cell(row=r, column=4).value = "推迟到达"
        ws1.cell(row=r, column=5).value = overload_way[key][4][0][0]
        ws1.cell(row=r, column=6).value = overload_way[key][4][0][1]
        r += 1

wb.save(filename=overload_file)