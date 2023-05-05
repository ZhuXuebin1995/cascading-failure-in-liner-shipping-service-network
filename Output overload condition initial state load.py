#找出直达路径中明显的绕远的节点对
#分航线去找！
#累计航程？
#扬程回城，最远端的港口，借此划分！！！！
#按照航线去除绕行的路径！
#绕行有那些表征？
#处于增长趋势
import sys
#在计算最短路的基础上，统计各个节点区间的访问次数
#coding:utf-8
import openpyxl #导入工作蒲模块
from openpyxl.utils import get_column_letter #导入模块中读取最大列的函数
from os import listdir
import os
import datetime
#设置的变量务必不可以跟模块、函数等重名！！！
import time
import copy
from geopy.distance import geodesic
import calendar
import json
from openpyxl import Workbook
start_time = time.time()
#===================================================================================================================
#为算法加载初步的数据
#将文件夹中的所有工作簿的工作表以邻接、直达的路径时间导入短期边中，边包含所属航线信息
#载入坐标以及经纬距离公式
aim_dress = r"E:\大连海事大学\博士\班轮运输时空复杂网络级联失效评估模型\地图\地图\港口坐标.xlsx"
wb_data = openpyxl.load_workbook(aim_dress)
active_sheet = wb_data['剔除过、清洗过0820']
row_number = active_sheet.max_row
xy_dict = {}
for i in range(2, row_number + 1):
    key = active_sheet.cell(i, 1).value
    x = active_sheet.cell(i, 2).value #维度
    y = active_sheet.cell(i, 3).value #精度
    xy_dict[key] = [x, y]


def count_distant(port_coordinate, point1, point2):
    x1 = port_coordinate[point1][0]
    y1 = port_coordinate[point1][1]
    x2 = port_coordinate[point2][0]
    y2 = port_coordinate[point2][1]
    distant_num = float(format(geodesic((x1, y1), (x2, y2)).km,'.2f')) #（维度，精度）
    return distant_num

#根据航线输出航线中相邻港口间的距离列表
def adjacent_length(aim_list,xy_message):
    adjacent_1 = []
    k = 0
    while k < len(aim_list) - 2:
        port_1 = aim_list[k]
        port_2 = aim_list[k+2]
        distance = count_distant(xy_message,port_1,port_2)
        adjacent_1.append(0)
        adjacent_1.append(distance)
        k += 2
    return adjacent_1

def cumulative_length(adjacent_1,index_1,index_2):#记录相邻距离的列表，对应的港口所在的位置1，位置2
    length = 0
    c_k = index_1
    while c_k < index_2:
        length += adjacent_1[c_k]
        c_k += 1
    return length



path_list = []
open_file_path = r'E:\大连海事大学\博士\数据\集装箱船公司船期数据\中远海集运（按照区域航线划分）\标准化数据'
file_name = listdir(open_file_path)
for i in file_name:
    com_path = os.path.join(open_file_path,i)
    path_list.append(com_path)
#print(path_list)
#path_list中已经有所有的工作蒲地址
port_list = [] #临时表
true_port_list = [] #永久列表
time_list = [] #临时表
date_list = {} #临时表日期表换成字典储存
datetime_list = {}#临时表日期时间变更为字典储存
whole_od_dict = {}
load_dict = {}#储存时刻使用次数的统计容器
l_num = 0
t_dict = {}
travel_time_dict = {}
line_dict = {}

def load_file(address):
    file_address = address
    with open(file_address) as add_json:
        new_dict = json.load(add_json)
    return new_dict
def build_file(address,aim_file):
    file_address = address
    trans_file = json.dumps(aim_file)
    with open(file_address,'w') as json_build_object:
        json_build_object.write(trans_file)

od_min_time = load_file('od min time.json')

port_in_line = {}

adjacent_list = []

#计算一边最短距离
for i in path_list:#遍历每一个工作蒲
    wb = openpyxl.load_workbook(i)
    sheet_names = wb.sheetnames # 获取工作蒲中所有表格
    #从表格中提取港口名、时刻、日期存入三个临时列表（重置）
    for j in sheet_names:
        strs = str(j) #按照航线名称创建字典键
        sheet = wb[j]
        max_column = sheet.max_column
        max_row = sheet.max_row #获取最大行数
        column = get_column_letter(max_column)
        row_1 = sheet["I3":"%s3" % column]
        for row1_cells in row_1:
            for cell in row1_cells:
                port_list.append(str(cell.value))#读取出每个工作表的港口信息
        adjacent_list = adjacent_length(port_list,xy_dict)
        row_2 = sheet["I5":"%s5" % column]
        for row2_cells in row_2:
            for cell in row2_cells:
                time_list.append(str(cell.value))
        #遍历行，对row3做修正==========================================================修正！！！！成立一个关日期行的字典？后期添加航线字典也使用字典值的遍历！！！！！
        #改变航线列表，后续所有与航线列表相关的代码也需要更改
        for row_number in range(6,max_row + 1):
            first_column = str("I"+str(row_number))
            last_column = str("%s" + str(row_number))
            row_3 = sheet[first_column:last_column % column]#row_3 = sheet["I6":"%s6" % column]
            voyage_number = len(date_list)
            date_list[voyage_number] = []
            for row3_cells in row_3:
                for cell in row3_cells:
                    date_list[voyage_number].append(str(cell.value))
                    #date_list,也需要变成字典
        # for i in date_list:
        #     print(str(i) + str(date_list[i]))
        #得到一个表的一个港口、时刻、日期列表
        #将日期、时刻列表合成为时刻信息标准列表new_datetime_list
        # print(j)
        # print(port_list)
        for port in port_list:
            if port not in true_port_list:
                if port != 'None':
                    true_port_list.append(port)#生成真实的港口列表
                    load_dict[port] = {}
                    #计划设计{}中的数据为time：[次数，装卸货时间]当前是到达时间，则是卸货，前时间-后时间符号为负的；当前是离开时间，则是装货，后时间-前时间，符号为正的
                    #不确定这样的赋值方式是否适用！！！！！！！！！！
        n = len(port_list)
        #print(str(n))
        #在检查结果的时候，出现"None 13:00"的字样，原因是有航次在这个日期没有船舶靠泊的安排，但是不能将其从列表中删除，否则出现列表串位的情况
        #因此在发现有None的时候，将关键位上的数据整个变成None值，后期计算时，做出处理即可!!!
        #date_list={0:[t1,t2..],1:[]...},修改datetime为字典属性
        for voyage_number in date_list.keys():
            k = 0
            datetime_list[voyage_number] = []
            while k < n:#这里并不是逐行添加的，注意！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！
                if date_list[voyage_number][k] != 'None':
                    datetime_number = str(date_list[voyage_number][k][:10]) + " " + str(time_list[k])
                else:
                    datetime_number = 'None'
                datetime_list[voyage_number].append(datetime_number)
                k += 1
        l_num += len(date_list)
        # for i in datetime_list:
        #     print(str(i) + str(datetime_list[i]))
        #value = get_time(0,2,datetime_list)
        #print(value)
        #======================================================================================
        #利用工作蒲中的数据为算法添加初始的邻接数据
        #采用p空间的方法存入邻接表数据，即在一条航线上就表示可以直达
        #航线是环形的，利用错位组合的方法不可避免的会造成不合理的长路径，要避免长路径的出现
        port_count = len(port_list)
        #对于出现列表中数据为空的情况，不添加对应的关键边
        #理想中的状态是按照航线分别储存航次信息
        #如果采用列表添加的模式，原本的上层正序、下层倒叙的数据读取方式不再合适
        k = 0
        for voyage_number in datetime_list.keys():
            p_num = 0
            l_n = str(j) + "." + str(voyage_number)
            while p_num < port_count:
                port_name = port_list[p_num]
                if datetime_list[voyage_number][p_num] != 'None':
                    try:
                        arrive_time = datetime_list[voyage_number][p_num]
                        arrive_time_1 = datetime.datetime.strptime(arrive_time, "%Y-%m-%d %H:%M")
                        leave_time = datetime_list[voyage_number][p_num + 1]
                        leave_time_1 = datetime.datetime.strptime(leave_time, "%Y-%m-%d %H:%M")
                    except ValueError:
                        print(str(i))
                        print(str(j))
                        print(str(voyage_number))
                    load_dict[port_name][(l_n, arrive_time_1)] = [0, leave_time_1 - arrive_time_1, -1]
                    load_dict[port_name][(l_n, leave_time_1)] = [0, leave_time_1 - arrive_time_1, 1]
                    # 对应的数据结构是，load_dict={港口名：{到达时间：[计数，卸货时间区间,负数]，离开时间：[计数，装货时间区间,正数】，。。。}}
                    # 更换一下初始的数据
                p_num += 2
            point_1 = 0
            while point_1 < port_count - 2:  # 点坐标2n，列表长度2n+1
                if datetime_list[voyage_number][point_1 + 1] != 'None':
                    start_point = port_list[point_1]
                    point_2 = port_count - 2
                    while point_2 > point_1:  # 船期中od信息只能是向后的,所以在起始点后一位结束
                        if datetime_list[voyage_number][point_2] != 'None':
                            end_point = port_list[point_2]
                            if start_point != end_point:  # 删除掉本地到本地的情况
                                depart_time_point = point_1 + 1
                                arrive_time_point = point_2
                                depart_time = datetime_list[voyage_number][depart_time_point]
                                arrive_time = datetime_list[voyage_number][arrive_time_point]
                                depart_time_1 = datetime.datetime.strptime(depart_time,"%Y-%m-%d %H:%M")
                                arrive_time_1 = datetime.datetime.strptime(arrive_time,"%Y-%m-%d %H:%M")
                                #改变line_dict
                                voyage = strs + "." + str(voyage_number)
                                # travel_time_dict[str(od)].append(travel_time)
                                #如果累计航程<3估计航程
                                cumulative_t = cumulative_length(adjacent_list,point_1,point_2)
                                estimate_t = count_distant(xy_dict,start_point,end_point)
                                if cumulative_t < 3 * estimate_t:
                                    line_dict[start_point,end_point] = [start_point,end_point,depart_time_1,arrive_time_1,voyage]
                                    load_dict[start_point][(l_n, depart_time_1)][0] += 1
                                    load_dict[end_point][(l_n, arrive_time_1)][0] += 1
                                #原本是航线字典={od:[]}
                                #将当前航线中的od信息导入:whole_od_list{(O,D):[[[1]],[[2]],[[3]]...]}，(O,D):[[1],[2]...].....}
                            point_2 -= 2
                        else:
                            point_2 -= 2
                    point_1 += 2
                else:
                    point_1 += 2
            for od in line_dict.keys():
                if od in whole_od_dict.keys():
                    whole_od_dict[od].append(copy.deepcopy([line_dict[od]]))
                else:
                    whole_od_dict[od] = []
                    whole_od_dict[od].append(copy.deepcopy([line_dict[od]]))
            line_dict.clear()
            # 定期清空临时表格实现数据导入，必不可少！！！
        datetime_list.clear()
        date_list.clear()
        port_list.clear()
        time_list.clear()
# build_file('new origin strain od time.json',travel_time_dict)


print("finished od strain_od")
print("add " + str(len(whole_od_dict)) + " od-pair")


j = 0
for key in whole_od_dict.keys():
    j += len(whole_od_dict[key])
print("or od length: " + str(j))
K = 0
for key in load_dict.keys():
    k += len(load_dict[key])
print("or load blanket: " + str(k))
#找出空的节点对
#
#至此所有数据均读取并且处理完毕！！！！！！！whole_od_dict ={od：[[[t1,t2，line]],[[t1,t2，line]]....}
t_num = len(true_port_list)
null_od_pair = []
o_num = 0
#k = 0 #检验节点对个数
#j = 0
while o_num < t_num:#按照起始节点，将节点对分类
    o_port = true_port_list[o_num]
    d_num = 0 #终点指针
    while d_num < t_num:
        d_port = true_port_list[d_num]
        if o_port != d_port:
            null_od_pair.append(tuple([o_port,d_port]))
            #j += 1
        d_num += 1
    o_num += 1
print("total od" + str(len(null_od_pair)))
#剔除已经直达的节点对
for od in whole_od_dict.keys():
    null_od_pair.remove(od)
print("after remove" + str(len(null_od_pair)))

# country_port = load_file('country od list.json')
# for c_od in country_port:
#     try:
#         null_od_pair.remove(tuple(c_od))
#     except ValueError:
#         continue
# print(len(null_od_pair))

def load_counter(in_list,count_can):#(待统计的路径列表，计数容器）
    #whole od dict = {od；[[[],[],[]],...]}
    for s_list in in_list:
        for od_path in s_list:
            p_1 = od_path[0]
            #[start_point,end_point,depart_time_1, arrive_time_1,voyage]
            p_1_leave = od_path[2]
            p_2 = od_path[1]
            p_2_arrive = od_path[3]
            l_m = od_path[-1]
            count_can[p_1][(l_m,p_1_leave)][0] += 1
            count_can[p_2][(l_m,p_2_arrive)][0] += 1
od_start = {}
od_end = {}
handing_time = datetime.timedelta(days=1)

for od_pair in whole_od_dict.keys():
    od_start[od_pair[0]] = []
    od_end[od_pair[1]] = []
for od_pair in whole_od_dict.keys():
    od_start[od_pair[0]].append(od_pair)
    od_end[od_pair[1]].append(od_pair)
print("there have " + str(len(null_od_pair)) + "od can't go strain")

def time_estimate(od_combination):
    sailing_time = 0
    for part in od_combination:
        sailing_time += (part[3] - part[2]).days + (part[3] - part[2]).seconds/86400
    return sailing_time
t_times = 0

while null_od_pair:
    check_1 = len(null_od_pair)
    for od in null_od_pair:
        k = 0
        temp_lists = [] #每遍历一个节点对，求出来的临时路径初始值都是空的
        voyage_check = {}#每遍历一个节点对，对应的船期检验数据列表都是空的
        for s_od in od_start[od[0]]:
            for e_od in od_end[od[1]]:
                if s_od[1] == e_od[0]:
                    for info_1 in whole_od_dict[s_od]:
                        temp_list = []
                        for info_2 in whole_od_dict[e_od]:
                            if info_1[-1][3] + handing_time < info_2[0][2]:
                                if temp_list:
                                    if info_2[-1][3] < temp_list[-1][3]:
                                        temp_list = info_2
                                else:
                                    temp_list = info_2
                        if temp_list:#在接续od对中找到最短路径后
                            new_list = info_1 + temp_list  #得到中转路径
                            v_id = info_1[0][4]
                            if v_id in voyage_check.keys():#确保之前没有相同的船期
                                #替换原本的路径！
                                c_num3 = voyage_check[v_id]
                                if new_list[-1][3] < temp_lists[c_num3][-1][3]:
                                    temp_lists[c_num3] = new_list
                            else:
                                voyage_check[v_id] = k
                                temp_lists.append(new_list)
                                k += 1
        if temp_lists:
            whole_od_dict[od] = copy.deepcopy(temp_lists)
            load_counter(temp_lists,load_dict)
            od_start[od[0]].append(od)
            od_end[od[1]].append(od)
            null_od_pair.remove(od)
    t_times += 1
    print("finish " + str(t_times) + " transport way finds")
    check_2 = len(null_od_pair)   #4644个节点对无法连通
    find_number = check_1 - check_2
    print("add " + str(find_number) + "od-pair")
    if find_number == 0 :
        print("there is still have no way od-pair")
        print(check_2)
        print(str(null_od_pair))
        break
print(len(whole_od_dict))

print("Finish whole od find!")

print("Finish load counter!")


end_time = time.time()
print("use " + str(end_time-start_time) + "seconds")
def build_interval_can(years):#按照某年生成一个区间树容器
    month_num = list(range(1, 13))#生成月份列表
    clock_list = list(range(0, 48))#生成时间区间列表
    times = dict.fromkeys(month_num, {})
    for m in times.keys():
        monthrange = calendar.monthrange(years, m)#按照某年的月份取一个月的天数
        d = monthrange[-1]#按照2020年的月份取一个月的天数
        day_list1 = list(range(1,d + 1)) #按照天数生成列表
        times[m] = dict.fromkeys(day_list1,{})
    for m in times.keys():
        for d in times[m].keys():
            times[m][d] = dict.fromkeys(clock_list)
            for t in times[m][d].keys():
                times[m][d][t] = {} #初始化为0
            #times[j][k] = dict.fromkeys(time_list,[0,0])
            #列表中0位置上存放装货负荷，1位上存放卸货负荷,2位上存放装货的航线船期，3位置上存放卸货的航线船期
    return times
'''生成区间树容器'''
#按照港口划分区间树，需要调整！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！



def count_load(load_dicts,port_interval_tree):#time_unit是累加的时间单元，默认为半小时
    for p in load_dicts.keys():
        times_can = build_interval_can(2020)
        for t in load_dicts[p].keys():
            #计算时间差有多少个半小时
            time_diff = load_dicts[p][t][1]
            half_num = time_diff.days*48 + time_diff.seconds/1800
            #计算实际负荷
            true_load = load_dicts[p][t][0]/half_num
            t_count = t[1]
            if load_dicts[p][t][2] < 0 :
                while t_count <= t[1] + time_diff:
                    # 按照访问时间，获取月份、天
                    month_num = t_count.month
                    date_num = t_count.day
                    hour_num = t_count.hour*2 + t_count.minute/30
                    #这个取大小，比较对象也有问题
                    if t[0] in times_can[month_num][date_num][hour_num]:
                        times_can[month_num][date_num][hour_num][t[0]] += true_load
                    else:
                        times_can[month_num][date_num][hour_num][t[0]] = true_load
                    t_count += datetime.timedelta(minutes=30)
                    #因为load_dict中的时间戳不是排序的，所以利用函数区间函数进行统计的时候，取最大可能是局部的最大
            else:
                t_count -= datetime.timedelta(minutes=30)
                while t_count >= t[1] - time_diff:
                    month_num = t_count.month
                    date_num = t_count.day
                    hour_num = t_count.hour * 2 + t_count.minute / 30
                    if t[0] in times_can[month_num][date_num][hour_num]:
                        times_can[month_num][date_num][hour_num][t[0]] += true_load
                    else:
                        times_can[month_num][date_num][hour_num][t[0]] = true_load
                    t_count -= datetime.timedelta(minutes=30)
        port_interval_tree[p] = copy.deepcopy(times_can)
        times_can.clear()

port_max_load = {}
port_interval_dict = {}
count_load(load_dict,port_interval_dict)

for i in range(1,9):
    t = 0
    overload_or_load = {}
    overload_file_address = "中断" + str(i) + "天实验\\overload way.json"
    overload_dict = load_file(overload_file_address)
    for key in overload_dict:
        p1 = overload_dict[key][0]
        time1 = datetime.datetime.strptime(overload_dict[key][1],"%Y-%m-%d %H:%M")
        m_num = time1.month
        d_num = time1.dayn
        h_num = time1.hour*2 + time1.minute/30
        overload_or_load[t] = []
        for l in port_interval_dict[p1][m_num][d_num][h_num]:
            overload_or_load[t].append([l,port_interval_dict[p1][m_num][d_num][h_num][l]])
        t += 1
    build_file_address = "中断" + str(i) + "天实验\\overload or load.json"
    build_file(build_file_address,overload_or_load)
    del overload_or_load


