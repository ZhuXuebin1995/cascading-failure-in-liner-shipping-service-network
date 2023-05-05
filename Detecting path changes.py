##找到发生变化的路
#发生了多少变化？
import json
import time
from openpyxl import Workbook
from multiprocessing import pool
def load_file(aim_address):
    with open(aim_address) as add_json:
        new_json = json.load(add_json)
    return new_json

def build_file(aim_address2,aim_file):
    trans_file = json.dumps(aim_file)
    with open(aim_address2,'w') as json_file:
        json_file.write(trans_file)


original_whole_od = load_file('whole od path (key as number).json')
for i in range(1,9):
    break_time = "中断" + str(i) + "天实验"
    aim_file_address = break_time + "\\whole od path (key as number).json"
    aim_file_address2 = break_time + "\\出发与到达差别.json"
    contrast_whole_od = load_file(aim_file_address)
    s_time = time.time()
    diff_dict = {}
    break_num = 0
    change_num1 = 0

    for key in original_whole_od.keys():
        k = 0
        while k < len(original_whole_od[key]):#确认过节点对总数不变
            if original_whole_od[key][k] != contrast_whole_od[key][k]:
                if original_whole_od[key][k][-1][1] == contrast_whole_od[key][k][-1][1]:
                    change_num1 += 1
                    if key in diff_dict.keys():
                        diff_dict[key].append([original_whole_od[key][k][0][2],original_whole_od[key][k][-1][3],
                                      contrast_whole_od[key][k][0][2],contrast_whole_od[key][k][-1][3]])
                    else:
                        diff_dict[key] = [[original_whole_od[key][k][0][2],original_whole_od[key][k][-1][3],
                                      contrast_whole_od[key][k][0][2],contrast_whole_od[key][k][-1][3]]]
                else:
                    break_num += 1
            k += 1

    workbook = Workbook()
    sheet = workbook.active
    workbook_address = break_time + '\\路径单元改变的数目与起始时间变化情况2.xlsx'
    sheet.cell(row=1,column=1).value = '中断的数目'

    sheet.cell(row=1,column=2).value = break_num
    sheet.cell(row=1, column=3).value = '改变的路径数目'

    sheet.cell(row=1, column=4).value = change_num1
    sheet.cell(row=2, column=1).value = "节点对编号"
    sheet.cell(row=2,column=2).value = "原始出发时间"
    sheet.cell(row=2, column=3).value = "原始到达时间"
    sheet.cell(row=2, column=4).value = "新出发时间"
    sheet.cell(row=2, column=5).value = "新到达时间"
    sheet.cell(row=2, column=6).value = "变化状态"
    num = 3
    change_num2 = 0
    for key in diff_dict.keys():
        change_num2 += len(diff_dict[key])
        for message in diff_dict[key]:

            if message[0] == message[2] and message[1] == message[3]:#期间改变，但是前后时间不受影响
                pass
            else:
                if message[0] != message[2] and message[1] == message[3]:
                    sheet.cell(row=num, column=1).value = key
                    sheet.cell(row=num,column=2).value = message[0]
                    sheet.cell(row=num, column=3).value = message[1]
                    sheet.cell(row=num, column=4).value = message[2]
                    sheet.cell(row=num, column=5).value = message[3]
                    sheet.cell(row=num, column=6).value = 0#出发时间改变
                elif message[0] == message[2] and message[1] != message[3]:
                    sheet.cell(row=num, column=1).value = key
                    sheet.cell(row=num, column=2).value = message[0]
                    sheet.cell(row=num, column=3).value = message[1]
                    sheet.cell(row=num, column=4).value = message[2]
                    sheet.cell(row=num, column=5).value = message[3]
                    sheet.cell(row=num, column=6).value = 1  # 到达时间改变
                elif message[0] != message[2] and message[1] != message[3]:
                    sheet.cell(row=num, column=1).value = key
                    sheet.cell(row=num, column=2).value = message[0]
                    sheet.cell(row=num, column=3).value = message[1]
                    sheet.cell(row=num, column=4).value = message[2]
                    sheet.cell(row=num, column=5).value = message[3]
                    sheet.cell(row=num, column=6).value = 2  # 出发到达时间改变
                num += 1
    sheet.cell(row=1, column=5).value = change_num2
    workbook.save(workbook_address)
    workbook.close()


e_time1 = time.time()
print(str(e_time1-s_time))

