#coding:utf-8
#break impact.json
#overload port original voyage.json
#overload voyage.json
import json
import openpyxl
from openpyxl import Workbook

xy_excel = openpyxl.load_workbook("港口坐标.xlsx")
xy_sheet = xy_excel['最终版']
row = xy_sheet.max_row
port_xy = {} #port:{纬度，经度}
for num in range(2,row+1):
    port = xy_sheet.cell(row=num,column=1).value
    port_y = xy_sheet.cell(row=num,column=2).value
    port_x = xy_sheet.cell(row=num,column=3).value
    port_xy[port] = [port_y,port_x]
xy_excel.close()

def load_json(file_address):
    with open(file_address) as json_object:
        aim_file = json.load(json_object)
    return aim_file
for i in range(1,9):
    break_impact = "中断" + str(i) + "天实验"
    break_impact_list = load_json(break_impact + "\\break impact.json") #[["PORT SAID", "AME3-1.1", 21600], ["SUEZ CANAL", "0AEU5.3", 50400], ["PORT SAID", "GEM2.9", 32400]]
    break_impact_dict = {}
    for k in break_impact_list: #生成由过载影响的船期列表
        break_impact_dict[k[1]] = [[k[0]],[]] #[[初始],[无中生有]]

    #过载时初始状态下的船期情况，
    overload_original_voyage = load_json(break_impact + "\\overload port original voyage.json")#[["AME2.5", "AME2.16", "CAX1.15", "WAX2-2.0", "MEX4-1.12", "MEX4-2.9", "AWE1.9", "AWE1.19", "CEN3.2", "MPNW1.9"], ["WAX2-2.0"], ["ESA.15"],



    overload_way_dict = load_json(break_impact + "\\overload or load.json")#{"0": ["QINGDAO", "2020-05-19 20:00", "2020-05-19 23:00", [], [["AWE1.19", 10800]]], "1": ["INCHEON"
    overload_dict = {}
    for key in overload_way_dict.keys():
        if overload_way_dict[key][3]:
            for v in overload_way_dict[key][3]:
                if v[0] in overload_dict:
                    overload_dict[v[0]][0].append(overload_way_dict[key][0])
                else:
                    overload_dict[v[0]] = [[overload_way_dict[key][0]],[]]#[[原有影响],[无中生有的影响]]
        if overload_way_dict[key][4]:
            for v in overload_way_dict[key][4]:
                if v[0] in overload_dict:
                    overload_dict[v[0]][0].append(overload_way_dict[key][0])
                else:
                    overload_dict[v[0]] = [[overload_way_dict[key][0]],[]]#[[原有影响],[无中生有的影响]]


    overload_voyage = load_json(break_impact + "\\overload voyage.json") #{"1": ["MEX4-2.9", "MPNW1.9", "WAX2-2.0", "CAX1.15", "CEN3.2", "MEX4-1.12", "AME2.5", "AME2.16", "AWE1.9", "AWE1.19"],

    #港口需要编号！
    change_dict = {}
    times = 0
    #需要一个单元，记录上一次船期所在港口，用初始的生成字典，建后面的内容用来更新迭代！！

    for key in overload_voyage.keys():
        port = overload_way_dict[str(times)][0]
        change_dict[times] = [port,[], [], []] #[港口，1中断影响，2过载影响，3其他影响]
        for v in overload_voyage[key]:
            if v in overload_original_voyage[times]:
                continue
            else:#与原本的船期情况不一致，找出不一致的原因，与两个变量库进行对比，[],[],[] -(中断影响，过载影响，无里头）上一个影响的港
                #不来自中断的影响，就是来自其他的影响
                if v in break_impact_dict.keys():
                    break_impact_dict[v][1].append(port)
                    change_dict[times][1].append(v)
                elif v in overload_dict.keys():
                    overload_dict[v][1].append(port)
                    change_dict[times][2].append(v)
                else:
                    change_dict[times][3].append(v)

        times += 1

    print("break impact dict")
    print(break_impact_dict)
    print("change dict")
    print(change_dict)
    print("overload dict")
    print(overload_dict)

    aim_xl = Workbook()
    Sheet1 = aim_xl.create_sheet("Break")

    Sheet1.cell(row=1,column=1).value = 'start_port'
    Sheet1.cell(row=1,column=2).value = 'y_1'
    Sheet1.cell(row=1,column=3).value = 'x_1'
    Sheet1.cell(row=1,column=4).value = 'end_port'
    Sheet1.cell(row=1,column=5).value = 'y_2'
    Sheet1.cell(row=1,column=6).value = 'x_2'
    Sheet1.cell(row=1,column=7).value = 'voyage'

    r = 2
    for key in break_impact_dict.keys():
        n = len(break_impact_dict[key][1])
        if n != 0:
            temp_list = break_impact_dict[key][0] + break_impact_dict[key][1]
            m = len(temp_list)
            k = 0
            while k < m - 1:
                p1 = temp_list[k]
                p2 = temp_list[k + 1]
                Sheet1.cell(row=r, column=1).value = p1
                Sheet1.cell(row=r,column=2).value = port_xy[p1][0]
                Sheet1.cell(row=r, column=3).value = port_xy[p1][1]
                Sheet1.cell(row=r, column=4).value = p2
                Sheet1.cell(row=r, column=5).value = port_xy[p2][0]
                Sheet1.cell(row=r, column=6).value = port_xy[p2][1]
                Sheet1.cell(row=r, column=7).value = key
                r += 1
                k += 1


    Sheet2 = aim_xl.create_sheet("overload")

    Sheet2.cell(row=1,column=1).value = 'start_port'
    Sheet2.cell(row=1,column=2).value = 'y_1'
    Sheet2.cell(row=1,column=3).value = 'x_1'
    Sheet2.cell(row=1,column=4).value = 'end_port'
    Sheet2.cell(row=1,column=5).value = 'y_2'
    Sheet2.cell(row=1,column=6).value = 'x_2'
    Sheet2.cell(row=1,column=7).value = 'voyage'

    r = 2
    for key in overload_dict.keys():
        n = len(overload_dict[key][1])
        if n != 0:
            temp_list = [overload_dict[key][0][-1]] + overload_dict[key][1]
            m = len(temp_list)
            k = 0
            while k < m - 1:
                p1 = temp_list[k]
                p2 = temp_list[k + 1]
                Sheet2.cell(row=r, column=1).value = p1
                Sheet2.cell(row=r,column=2).value = port_xy[p1][0]
                Sheet2.cell(row=r, column=3).value = port_xy[p1][1]
                Sheet2.cell(row=r, column=4).value = p2
                Sheet2.cell(row=r, column=5).value = port_xy[p2][0]
                Sheet2.cell(row=r, column=6).value = port_xy[p2][1]
                Sheet2.cell(row=r, column=7).value = key
                r += 1
                k += 1

    Sheet3 = aim_xl.create_sheet("summary ")
    Sheet3.cell(row=1,column=1).value = "total overload"
    Sheet3.cell(row=1,column=2).value = len(change_dict)
    b_impact_num = 0
    o_impact_num = 0
    others = 0
    for key in change_dict.keys():
        if len(change_dict[key][1]) != 0:
            b_impact_num += 1
        if len(change_dict[key][2]) != 0:
            o_impact_num += 1
        if len(change_dict[key][3]) != 0:
            others += 1
    Sheet3.cell(row=2,column=1).value = "break"
    Sheet3.cell(row=2,column=2).value = b_impact_num
    Sheet3.cell(row=3,column=1).value = "overload"
    Sheet3.cell(row=3,column=2).value = o_impact_num
    Sheet3.cell(row=4,column=1).value = "others"
    Sheet3.cell(row=4,column=2).value = others


    aim_xl_address = break_impact + "\\" + break_impact + "过载链路传播.xlsx"
    aim_xl.save(aim_xl_address)
    aim_xl.close()

